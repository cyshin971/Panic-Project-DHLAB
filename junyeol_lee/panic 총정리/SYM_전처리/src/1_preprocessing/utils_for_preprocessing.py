import pandas as pd
import numpy as np
from openpyxl import load_workbook
from itertools import islice

def load_raw_file(paths, sheet_name):
    """
    단일 파일 경로(str) 또는 파일 경로 리스트(list)를 받아,
    지정된 시트를 DataFrame으로 불러온 뒤 모두 합쳐서 반환합니다.
    """
    if isinstance(paths, str):
        paths = [paths]

    all_dfs = []
    for path in paths:
        # 파일 확장자가 .xlsx, .xlsm 등 openpyxl 지원 형식인지 확인
        if not (path.endswith(".xlsx") or path.endswith(".xlsm") or path.endswith(".xltx") or path.endswith(".xltm")):
            raise ValueError(f"Invalid file format: {path!r}. .xlsx/.xlsm/.xltx/.xltm만 지원합니다.")
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet_name]
        data = ws.values
        cols = next(data)[0:]
        rows = list(data)
        _ = [r[0] for r in rows]  # 첫 열값을 인덱스로 읽었으나 실제로 사용하지 않음
        rows = (islice(r, 0, None) for r in rows)
        df = pd.DataFrame(rows, columns=cols)
        df.replace("", np.nan, inplace=True)
        df.columns = [
            col.replace(" ", "_").replace("-", "_") if col is not None else col
            for col in df.columns
        ]
        all_dfs.append(df)

    return pd.concat(all_dfs, ignore_index=True)

def serialize_lifelog_heartrate(path):
    """
    '라이프로그-심박수' 시트를 불러와 1분 단위 롱 포맷(날짜, ID, 시간, 심박수)으로 변환합니다.
    
    반환 컬럼:
    - date (날짜)
    - ID (비식별키)
    - time (HH:MM:SS)
    - heart_rate (각 분 단위 심박수, -1은 결측값)
    """
    # 1) 원본 시트 로드
    df = load_raw_file(path, sheet_name='라이프로그-심박수')
    
    # 2) 컬럼명 한글 → 영문(관습적인 이름)으로 변경 (동적으로 컬럼명 탐지)
    # - '측정값'이 포함된 컬럼: values
    # - '비식별키' 포함: ID
    # - '날짜' 포함: date
    # - '평균' 포함: avg_heart_rate (있을 경우)
    value_col = next(col for col in df.columns if '측정값' in col)
    id_col = next(col for col in df.columns if '비식별키' in col)
    date_col = next(col for col in df.columns if '날짜' in col)
    # 평균 심박수는 있을 수도, 없을 수도 있음
    avg_col = next((col for col in df.columns if '평균' in col), None)
    rename_dict = {
        id_col: 'ID',
        date_col: 'date',
        value_col: 'values',
    }
    if avg_col:
        rename_dict[avg_col] = 'avg_heart_rate'
    df.rename(columns=rename_dict, inplace=True)
    
    # 2.1) 같은 date, ID에 중복된 행이 있을 경우 'values' 문자열을 이어붙이기
    agg_dict = {'values': lambda x: ','.join(x.dropna().astype(str))}
    if 'avg_heart_rate' in df.columns:
        agg_dict['avg_heart_rate'] = 'mean'
    df = df.groupby(['date', 'ID'], as_index=False).agg(agg_dict)
    
    # 3) 1분 단위 시간 인덱스 생성 ('00:00:00' ~ '23:59:00' 총 1440개)
    temp_idx = pd.date_range('2022-01-01', periods=1440, freq='1T')
    col_minutes = temp_idx.strftime('%H:%M:%S')
    
    # 4) 쉼표로 이어진 문자열을 1분 단위 컬럼으로 분리
    # values 열의 문자열이 "79,79,83,..." 형태로 들어있다고 가정
    df_splitted = pd.DataFrame(
        df['values'].str.split(',').tolist(),
        columns=col_minutes
    )
    df_splitted.reset_index(drop=True, inplace=True)
    
    # 5) 원본 df와 분리된 df_splitted를 인덱스 기준으로 병합
    df_reset = df.reset_index(drop=True)
    df_merged = pd.concat([df_reset[['date', 'ID']], df_splitted], axis=1)
    
    # 6) melt를 통해 롱 포맷으로 변환
    df_melted = df_merged.melt(id_vars=['date', 'ID'], var_name='time', value_name='heart_rate')
    
    # 7) heart_rate 컬럼을 숫자형으로 변환 (-1은 결측으로 처리)
    df_melted['heart_rate'] = pd.to_numeric(df_melted['heart_rate'], errors='coerce')
    df_melted.loc[df_melted['heart_rate'] == -1, 'heart_rate'] = np.nan
    
    # 8) 필요한 컬럼만 반환
    return df_melted[['date', 'ID', 'time', 'heart_rate']]



raw = """PXPN_10006
PXPN_10007
PXPN_10008
PXPN_10009
PXPN_10010
PXPN_10011
PXPN_10012
PXPN_10013
PXPN_10014
PXPN_10015
PXPN_10018
PXPN_10019
PXPN_10020
PXPN_10021
PXPN_10022
PXPN_10023
PXPN_10024
PXPN_10025
PXPN_10028
PXPN_10029
PXPN_10030
PXPN_10032
PXPN_10034
PXPN_10035
PXPN_10037
PXPN_10038
PXPN_10039
PXPN_10040
SYM1-1-100
SYM1-1-102
SYM1-1-103
SYM1-1-106
SYM1-1-107
SYM1-1-108
SYM1-1-109
SYM1-1-110
SYM1-1-111
SYM1-1-112
SYM1-1-113
SYM1-1-114
SYM1-1-115
SYM1-1-116
SYM1-1-117
SYM1-1-118
SYM1-1-119
SYM1-1-120
SYM1-1-121
SYM1-1-122
SYM1-1-123
SYM1-1-124
SYM1-1-125
SYM1-1-126
SYM1-1-127
SYM1-1-128
SYM1-1-129
SYM1-1-130
SYM1-1-131
SYM1-1-132
SYM1-1-133
SYM1-1-134
SYM1-1-136
SYM1-1-139
SYM1-1-142
SYM1-1-145
SYM1-1-146
SYM1-1-147
SYM1-1-148
SYM1-1-149
SYM1-1-150
SYM1-1-151
SYM1-1-152
SYM1-1-153
SYM1-1-154
SYM1-1-155
SYM1-1-156
SYM1-1-157
SYM1-1-158
SYM1-1-159
SYM1-1-160
SYM1-1-161
SYM1-1-162
SYM1-1-163
SYM1-1-164
SYM1-1-165
SYM1-1-166
SYM1-1-167
SYM1-1-168
SYM1-1-169
SYM1-1-170
SYM1-1-171
SYM1-1-172
SYM1-1-173
SYM1-1-174
SYM1-1-175
SYM1-1-176
SYM1-1-177
SYM1-1-178
SYM1-1-179
SYM1-1-180
SYM1-1-181
SYM1-1-182
SYM1-1-183
SYM1-1-184
SYM1-1-185
SYM1-1-186
SYM1-1-187
SYM1-1-188
SYM1-1-189
SYM1-1-190
SYM1-1-191
SYM1-1-192
SYM1-1-193
SYM1-1-194
SYM1-1-195
SYM1-1-196
SYM1-1-197
SYM1-1-198
SYM1-1-199
SYM1-1-200
SYM1-1-201
SYM1-1-202
SYM1-1-203
SYM1-1-204
SYM1-1-205
SYM1-1-206
SYM1-1-207
SYM1-1-208
SYM1-1-209
SYM1-1-210
SYM1-1-211
SYM1-1-212
SYM1-1-213
SYM1-1-214
SYM1-1-215
SYM1-1-216
SYM1-1-217
SYM1-1-218
SYM1-1-219
SYM1-1-220
SYM1-1-221
SYM1-1-222
SYM1-1-223
SYM1-1-224
SYM1-1-225
SYM1-1-226
SYM1-1-227
SYM1-1-228
SYM1-1-229
SYM1-1-230
SYM1-1-231
SYM1-1-232
SYM1-1-233
SYM1-1-235
SYM1-1-236
SYM1-1-237
SYM1-1-238
SYM1-1-239
SYM1-1-240
SYM1-1-241
SYM1-1-242
SYM1-1-243
SYM1-1-244
SYM1-1-245
SYM1-1-247
SYM1-1-248
SYM1-1-249
SYM1-1-250
SYM1-1-251
SYM1-1-252
SYM1-1-253
SYM1-1-254
SYM1-1-255
SYM1-1-256
SYM1-1-257
SYM1-1-258
SYM1-1-259
SYM1-1-260
SYM1-1-261
SYM1-1-262
SYM1-1-263
SYM1-1-264
SYM1-1-265
SYM1-1-266
SYM1-1-267
SYM1-1-268
SYM1-1-269
SYM1-1-270
SYM1-1-271
SYM1-1-273
SYM1-1-276
SYM1-1-277
SYM1-1-279
SYM1-1-280
SYM1-1-281
SYM1-1-282
SYM1-1-284
SYM1-1-286
SYM1-1-287
SYM1-1-289
SYM1-1-290
SYM1-1-291
SYM1-1-293
SYM1-1-294
SYM1-1-295
SYM1-1-296
SYM1-1-298
SYM1-1-299
SYM1-1-300
SYM1-1-302
SYM1-1-303
SYM1-1-304
SYM1-1-305
SYM1-1-306
SYM1-1-307
SYM1-1-308
SYM1-1-310
SYM1-1-311
SYM1-1-312
SYM1-1-313
SYM1-1-314
SYM1-1-315
SYM1-1-317
SYM1-1-318
SYM1-1-319
SYM1-1-320
SYM1-1-321
SYM1-1-322
SYM1-1-323
SYM1-1-324
SYM1-1-326
SYM1-1-328
SYM1-1-329
SYM1-1-331
SYM1-1-332
SYM1-1-333
SYM1-1-336
SYM1-1-338
SYM1-1-339
SYM1-1-340
SYM1-1-342
SYM1-1-343
SYM1-1-345
SYM1-1-346
SYM1-1-348
SYM1-1-350
SYM1-1-351
SYM1-1-352
SYM1-1-354
SYM1-1-356
SYM1-1-358
SYM1-1-359
SYM1-1-360
SYM1-1-361
SYM1-1-362
SYM1-1-364
SYM1-1-365
SYM1-1-367
SYM1-1-368
SYM1-1-369
SYM1-1-371
SYM1-1-373
SYM1-1-374
SYM1-1-375
SYM1-1-376
SYM1-1-377
SYM1-1-378
SYM1-1-379
SYM1-1-380
SYM1-1-382
SYM1-1-383
SYM1-1-385
SYM1-1-386
SYM1-1-387
SYM1-1-388
SYM1-1-389
SYM1-1-390
SYM1-1-391
SYM1-1-392
SYM1-1-394
SYM1-1-395
SYM1-1-397
SYM1-1-398
SYM1-1-400
SYM1-1-402
SYM1-1-404
SYM1-1-406
SYM1-1-407
SYM1-1-408
SYM1-1-410
SYM1-1-411
SYM1-1-413
SYM1-1-414
SYM1-1-415
SYM1-1-416
SYM1-1-417
SYM1-1-418
SYM1-1-419
SYM1-1-421
SYM1-1-423
SYM1-1-424
SYM1-1-426
SYM1-1-427
SYM1-1-428
SYM1-1-429
SYM1-1-430
SYM1-1-431
SYM1-1-432
SYM1-1-433
SYM1-1-434
SYM1-1-435
SYM1-1-436
SYM1-1-437
SYM1-1-439
SYM1-1-440
SYM1-1-441
SYM1-1-442
SYM1-1-443
SYM1-1-444
SYM1-1-446
SYM1-1-447
SYM1-1-449
SYM1-1-451
SYM1-1-452
SYM1-1-453
SYM1-1-454
SYM1-1-455
SYM1-1-456
SYM1-1-457
SYM1-1-458
SYM1-1-459
SYM1-1-460
SYM1-1-461
SYM1-1-462
SYM1-1-463
SYM1-1-464
SYM1-1-465
SYM1-1-466
SYM1-1-467
SYM1-1-468
SYM1-1-469
SYM1-1-470
SYM1-1-475
SYM1-1-479
SYM1-1-480
SYM1-1-481
SYM1-1-65
SYM1-1-66
SYM1-1-67
SYM1-1-68
SYM1-1-69
SYM1-1-70
SYM1-1-71
SYM1-1-74
SYM1-1-75
SYM1-1-77
SYM1-1-78
SYM1-1-79
SYM1-1-80
SYM1-1-81
SYM1-1-82
SYM1-1-83
SYM1-1-84
SYM1-1-85
SYM1-1-86
SYM1-1-88
SYM1-1-89
SYM1-1-90
SYM1-1-91
SYM1-1-93
SYM1-1-94
SYM1-1-95
SYM1-1-99
SYM2-1-101
SYM2-1-135
SYM2-1-137
SYM2-1-143
SYM2-1-246
SYM2-1-275
SYM2-1-278
SYM2-1-283
SYM2-1-285
SYM2-1-288
SYM2-1-292
SYM2-1-297
SYM2-1-301
SYM2-1-309
SYM2-1-316
SYM2-1-325
SYM2-1-327
SYM2-1-330
SYM2-1-334
SYM2-1-337
SYM2-1-341
SYM2-1-344
SYM2-1-347
SYM2-1-349
SYM2-1-353
SYM2-1-355
SYM2-1-357
SYM2-1-363
SYM2-1-366
SYM2-1-370
SYM2-1-372
SYM2-1-381
SYM2-1-384
SYM2-1-393
SYM2-1-396
SYM2-1-399
SYM2-1-401
SYM2-1-403
SYM2-1-405
SYM2-1-409
SYM2-1-420
SYM2-1-422
SYM2-1-438
SYM2-1-445
SYM2-1-448
SYM2-1-471
SYM2-1-472
SYM2-1-473
SYM2-1-474
SYM2-1-476
SYM2-1-477
SYM2-1-478
SYM2-1-482
SYM2-1-62
SYM2-1-73
SYM2-1-96""".strip()
valid_ids = raw.split("\n")
# ------------------------------------------------------------
# 3) 범용 필터 함수
# ------------------------------------------------------------
def filter_by_valid_ids(df, id_column="ID"):
    """
    DataFrame(df)의 id_column 컬럼 값을 valid_ids 리스트에 포함된 ID에 한해 필터링한 뒤
    복사본을 반환합니다.
    
    예:
        df_filtered = filter_by_valid_ids(original_df, id_column="비식별키")
    """
    return df[df[id_column].isin(valid_ids)].copy()

# 함수 사용 예시
# file_path = '/mnt/data/backup_SYM1.xlsx'
# df_serialized_hr = serialize_lifelog_heartrate(file_path)
# print(df_serialized_hr.head())


import pandas as pd
import numpy as np
from openpyxl import load_workbook
from itertools import islice

def load_raw_file(paths, sheet_name):
    """
    단일 파일 경로(str) 또는 파일 경로 리스트(list)를 받아,
    지정된 시트를 DataFrame으로 불러온 뒤 모두 합쳐서 반환합니다.
    """
    if isinstance(paths, str):
        paths = [paths]

    all_dfs = []
    for path in paths:
        # 파일 확장자가 .xlsx, .xlsm 등 openpyxl 지원 형식인지 확인
        if not (path.endswith(".xlsx") or path.endswith(".xlsm") or path.endswith(".xltx") or path.endswith(".xltm")):
            raise ValueError(f"Invalid file format: {path!r}. .xlsx/.xlsm/.xltx/.xltm만 지원합니다.")
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet_name]
        data = ws.values
        cols = next(data)[0:]
        rows = list(data)
        _ = [r[0] for r in rows]  # 첫 열값을 인덱스로 읽었으나 실제로 사용하지 않음
        rows = (islice(r, 0, None) for r in rows)
        df = pd.DataFrame(rows, columns=cols)
        df.replace("", np.nan, inplace=True)
        df.columns = [
            col.replace(" ", "_").replace("-", "_") if col is not None else col
            for col in df.columns
        ]
        all_dfs.append(df)

    return pd.concat(all_dfs, ignore_index=True)

def serialize_lifelog_heartrate(path):
    """
    '라이프로그-심박수' 시트를 불러와 1분 단위 롱 포맷(날짜, ID, 시간, 심박수)으로 변환합니다.
    
    반환 컬럼:
    - date (날짜)
    - ID (비식별키)
    - time (HH:MM:SS)
    - heart_rate (각 분 단위 심박수, -1은 결측값)
    """
    # 1) 원본 시트 로드
    df = load_raw_file(path, sheet_name='라이프로그-심박수')
    
    # 2) 컬럼명 한글 → 영문(관습적인 이름)으로 변경 (동적으로 컬럼명 탐지)
    # - '측정값'이 포함된 컬럼: values
    # - '비식별키' 포함: ID
    # - '날짜' 포함: date
    # - '평균' 포함: avg_heart_rate (있을 경우)
    value_col = next(col for col in df.columns if '측정값' in col)
    id_col = next(col for col in df.columns if '비식별키' in col)
    date_col = next(col for col in df.columns if '날짜' in col)
    # 평균 심박수는 있을 수도, 없을 수도 있음
    avg_col = next((col for col in df.columns if '평균' in col), None)
    rename_dict = {
        id_col: 'ID',
        date_col: 'date',
        value_col: 'values',
    }
    if avg_col:
        rename_dict[avg_col] = 'avg_heart_rate'
    df.rename(columns=rename_dict, inplace=True)
    
    # 3) 1분 단위 시간 인덱스 생성 ('00:00:00' ~ '23:59:00' 총 1440개)
    temp_idx = pd.date_range('2022-01-01', periods=1440, freq='1T')
    col_minutes = temp_idx.strftime('%H:%M:%S')
    
    # 4) 쉼표로 이어진 문자열을 1분 단위 컬럼으로 분리
    # values 열의 문자열이 "79,79,83,..." 형태로 들어있다고 가정
    df_splitted = pd.DataFrame(
        df['values'].str.split(',').tolist(),
        columns=col_minutes
    )
    df_splitted.reset_index(drop=True, inplace=True)
    
    # 5) 원본 df와 분리된 df_splitted를 인덱스 기준으로 병합
    df_reset = df.reset_index(drop=True)
    df_merged = pd.concat([df_reset[['date', 'ID']], df_splitted], axis=1)
    
    # 6) melt를 통해 롱 포맷으로 변환
    df_melted = df_merged.melt(id_vars=['date', 'ID'], var_name='time', value_name='heart_rate')
    
    # 7) heart_rate 컬럼을 숫자형으로 변환 (-1은 결측으로 처리)
    df_melted['heart_rate'] = pd.to_numeric(df_melted['heart_rate'], errors='coerce')
    df_melted.loc[df_melted['heart_rate'] == -1, 'heart_rate'] = np.nan
    
    # 8) 필요한 컬럼만 반환
    return df_melted[['date', 'ID', 'time', 'heart_rate']]



id_text = """PXPN_10006
PXPN_10007
PXPN_10008
PXPN_10009
PXPN_10010
PXPN_10011
PXPN_10012
PXPN_10013
PXPN_10014
PXPN_10015
PXPN_10018
PXPN_10019
PXPN_10020
PXPN_10021
PXPN_10022
PXPN_10023
PXPN_10024
PXPN_10025
PXPN_10028
PXPN_10029
PXPN_10030
PXPN_10032
PXPN_10034
PXPN_10035
PXPN_10037
PXPN_10038
PXPN_10039
PXPN_10040
SYM1-1-100
SYM1-1-102
SYM1-1-103
SYM1-1-106
SYM1-1-107
SYM1-1-108
SYM1-1-109
SYM1-1-110
SYM1-1-111
SYM1-1-112
SYM1-1-113
SYM1-1-114
SYM1-1-115
SYM1-1-116
SYM1-1-117
SYM1-1-118
SYM1-1-119
SYM1-1-120
SYM1-1-121
SYM1-1-122
SYM1-1-123
SYM1-1-124
SYM1-1-125
SYM1-1-126
SYM1-1-127
SYM1-1-128
SYM1-1-129
SYM1-1-130
SYM1-1-131
SYM1-1-132
SYM1-1-133
SYM1-1-134
SYM1-1-136
SYM1-1-139
SYM1-1-142
SYM1-1-145
SYM1-1-146
SYM1-1-147
SYM1-1-148
SYM1-1-149
SYM1-1-150
SYM1-1-151
SYM1-1-152
SYM1-1-153
SYM1-1-154
SYM1-1-155
SYM1-1-156
SYM1-1-157
SYM1-1-158
SYM1-1-159
SYM1-1-160
SYM1-1-161
SYM1-1-162
SYM1-1-163
SYM1-1-164
SYM1-1-165
SYM1-1-166
SYM1-1-167
SYM1-1-168
SYM1-1-169
SYM1-1-170
SYM1-1-171
SYM1-1-172
SYM1-1-173
SYM1-1-174
SYM1-1-175
SYM1-1-176
SYM1-1-177
SYM1-1-178
SYM1-1-179
SYM1-1-180
SYM1-1-181
SYM1-1-182
SYM1-1-183
SYM1-1-184
SYM1-1-185
SYM1-1-186
SYM1-1-187
SYM1-1-188
SYM1-1-189
SYM1-1-190
SYM1-1-191
SYM1-1-192
SYM1-1-193
SYM1-1-194
SYM1-1-195
SYM1-1-196
SYM1-1-197
SYM1-1-198
SYM1-1-199
SYM1-1-200
SYM1-1-201
SYM1-1-202
SYM1-1-203
SYM1-1-204
SYM1-1-205
SYM1-1-206
SYM1-1-207
SYM1-1-208
SYM1-1-209
SYM1-1-210
SYM1-1-211
SYM1-1-212
SYM1-1-213
SYM1-1-214
SYM1-1-215
SYM1-1-216
SYM1-1-217
SYM1-1-218
SYM1-1-219
SYM1-1-220
SYM1-1-221
SYM1-1-222
SYM1-1-223
SYM1-1-224
SYM1-1-225
SYM1-1-226
SYM1-1-227
SYM1-1-228
SYM1-1-229
SYM1-1-230
SYM1-1-231
SYM1-1-232
SYM1-1-233
SYM1-1-235
SYM1-1-236
SYM1-1-237
SYM1-1-238
SYM1-1-239
SYM1-1-240
SYM1-1-241
SYM1-1-242
SYM1-1-243
SYM1-1-244
SYM1-1-245
SYM1-1-247
SYM1-1-248
SYM1-1-249
SYM1-1-250
SYM1-1-251
SYM1-1-252
SYM1-1-253
SYM1-1-254
SYM1-1-255
SYM1-1-256
SYM1-1-257
SYM1-1-258
SYM1-1-259
SYM1-1-260
SYM1-1-261
SYM1-1-262
SYM1-1-263
SYM1-1-264
SYM1-1-265
SYM1-1-266
SYM1-1-267
SYM1-1-268
SYM1-1-269
SYM1-1-270
SYM1-1-271
SYM1-1-273
SYM1-1-276
SYM1-1-277
SYM1-1-279
SYM1-1-280
SYM1-1-281
SYM1-1-282
SYM1-1-284
SYM1-1-286
SYM1-1-287
SYM1-1-289
SYM1-1-290
SYM1-1-291
SYM1-1-293
SYM1-1-294
SYM1-1-295
SYM1-1-296
SYM1-1-298
SYM1-1-299
SYM1-1-300
SYM1-1-302
SYM1-1-303
SYM1-1-304
SYM1-1-305
SYM1-1-306
SYM1-1-307
SYM1-1-308
SYM1-1-310
SYM1-1-311
SYM1-1-312
SYM1-1-313
SYM1-1-314
SYM1-1-315
SYM1-1-317
SYM1-1-318
SYM1-1-319
SYM1-1-320
SYM1-1-321
SYM1-1-322
SYM1-1-323
SYM1-1-324
SYM1-1-326
SYM1-1-328
SYM1-1-329
SYM1-1-331
SYM1-1-332
SYM1-1-333
SYM1-1-336
SYM1-1-338
SYM1-1-339
SYM1-1-340
SYM1-1-342
SYM1-1-343
SYM1-1-345
SYM1-1-346
SYM1-1-348
SYM1-1-350
SYM1-1-351
SYM1-1-352
SYM1-1-354
SYM1-1-356
SYM1-1-358
SYM1-1-359
SYM1-1-360
SYM1-1-361
SYM1-1-362
SYM1-1-364
SYM1-1-365
SYM1-1-367
SYM1-1-368
SYM1-1-369
SYM1-1-371
SYM1-1-373
SYM1-1-374
SYM1-1-375
SYM1-1-376
SYM1-1-377
SYM1-1-378
SYM1-1-379
SYM1-1-380
SYM1-1-382
SYM1-1-383
SYM1-1-385
SYM1-1-386
SYM1-1-387
SYM1-1-388
SYM1-1-389
SYM1-1-390
SYM1-1-391
SYM1-1-392
SYM1-1-394
SYM1-1-395
SYM1-1-397
SYM1-1-398
SYM1-1-400
SYM1-1-402
SYM1-1-404
SYM1-1-406
SYM1-1-407
SYM1-1-408
SYM1-1-410
SYM1-1-411
SYM1-1-413
SYM1-1-414
SYM1-1-415
SYM1-1-416
SYM1-1-417
SYM1-1-418
SYM1-1-419
SYM1-1-421
SYM1-1-423
SYM1-1-424
SYM1-1-426
SYM1-1-427
SYM1-1-428
SYM1-1-429
SYM1-1-430
SYM1-1-431
SYM1-1-432
SYM1-1-433
SYM1-1-434
SYM1-1-435
SYM1-1-436
SYM1-1-437
SYM1-1-439
SYM1-1-440
SYM1-1-441
SYM1-1-442
SYM1-1-443
SYM1-1-444
SYM1-1-446
SYM1-1-447
SYM1-1-449
SYM1-1-451
SYM1-1-452
SYM1-1-453
SYM1-1-454
SYM1-1-455
SYM1-1-456
SYM1-1-457
SYM1-1-458
SYM1-1-459
SYM1-1-460
SYM1-1-461
SYM1-1-462
SYM1-1-463
SYM1-1-464
SYM1-1-465
SYM1-1-466
SYM1-1-467
SYM1-1-468
SYM1-1-469
SYM1-1-470
SYM1-1-475
SYM1-1-479
SYM1-1-480
SYM1-1-481
SYM1-1-65
SYM1-1-66
SYM1-1-67
SYM1-1-68
SYM1-1-69
SYM1-1-70
SYM1-1-71
SYM1-1-74
SYM1-1-75
SYM1-1-77
SYM1-1-78
SYM1-1-79
SYM1-1-80
SYM1-1-81
SYM1-1-82
SYM1-1-83
SYM1-1-84
SYM1-1-85
SYM1-1-86
SYM1-1-88
SYM1-1-89
SYM1-1-90
SYM1-1-91
SYM1-1-93
SYM1-1-94
SYM1-1-95
SYM1-1-99
SYM2-1-101
SYM2-1-135
SYM2-1-137
SYM2-1-143
SYM2-1-246
SYM2-1-275
SYM2-1-278
SYM2-1-283
SYM2-1-285
SYM2-1-288
SYM2-1-292
SYM2-1-297
SYM2-1-301
SYM2-1-309
SYM2-1-316
SYM2-1-325
SYM2-1-327
SYM2-1-330
SYM2-1-334
SYM2-1-337
SYM2-1-341
SYM2-1-344
SYM2-1-347
SYM2-1-349
SYM2-1-353
SYM2-1-355
SYM2-1-357
SYM2-1-363
SYM2-1-366
SYM2-1-370
SYM2-1-372
SYM2-1-381
SYM2-1-384
SYM2-1-393
SYM2-1-396
SYM2-1-399
SYM2-1-401
SYM2-1-403
SYM2-1-405
SYM2-1-409
SYM2-1-420
SYM2-1-422
SYM2-1-438
SYM2-1-445
SYM2-1-448
SYM2-1-471
SYM2-1-472
SYM2-1-473
SYM2-1-474
SYM2-1-476
SYM2-1-477
SYM2-1-478
SYM2-1-482
SYM2-1-62
SYM2-1-73
SYM2-1-96"""
valid_ids = id_text.strip().splitlines()

# ------------------------------------------------------------
# 3) 범용 필터 함수
# ------------------------------------------------------------
def filter_by_valid_ids(df, id_column="ID"):
    """
    DataFrame(df)의 id_column 컬럼 값을 valid_ids 리스트에 포함된 ID에 한해 필터링한 뒤
    복사본을 반환합니다.
    
    예:
        df_filtered = filter_by_valid_ids(original_df, id_column="비식별키")
    """
    return df[df[id_column].isin(valid_ids)].copy()

# 함수 사용 예시
# file_path = '/mnt/data/backup_SYM1.xlsx'
# df_serialized_hr = serialize_lifelog_heartrate(file_path)
# print(df_serialized_hr.head())


def extract_emotion_diary_from_raw(path, questionnaire_sheet, df_name, questionnaire_column):
    df = load_raw_file(path, sheet_name=questionnaire_sheet)

    # Dynamically detect ID and date columns
    id_col = next((col for col in df.columns if ('비식별키' in col) or ('Non_identifying' in col)), None)
    date_col = next((col for col in df.columns if ('날짜' in col) or ('Date' in col)), None)

    if id_col is None:
        raise ValueError("Could not find ID column (비식별키 or Non_identifying) in columns: " + str(df.columns))
    if date_col is None:
        raise ValueError("Could not find date column (날짜 or Date) in columns: " + str(df.columns))
    if questionnaire_column not in df.columns:
        raise ValueError(f"Column '{questionnaire_column}' not found in DataFrame columns: {df.columns}")

    df = df[[id_col, date_col, questionnaire_column]]
    df.columns = ["ID", "date", df_name]
    df.drop_duplicates(['ID', 'date'], keep='last', inplace=True, ignore_index=False)
    return df

def extract_questionnaire_from_raw(path, questionnaire_sheet, df_name ,questionnaire_column):
    df = load_raw_file(path, sheet_name=questionnaire_sheet)
    df = df[['비식별키', '설문종료일', '설문완료일', questionnaire_column]]
    df.drop(['설문종료일'], axis=1, inplace=True)
    df.columns = ["ID", 'date', df_name]
    df.drop_duplicates(['ID','date'], keep='last', inplace=True, ignore_index=False)
    return df